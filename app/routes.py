from flask import render_template, flash, redirect, url_for, request, jsonify, send_file
from app import app, db, login_manager
from app.forms import LoginForm
from openpyxl import Workbook
from app.models import User, Contract, Service, Expense
from werkzeug.security import check_password_hash
from flask_login import login_user, login_required, logout_user, current_user
from functools import wraps
import datetime, os, random, string, zipfile, tempfile, requests



ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# Проверка расширения файла
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Функция для генерации уникального 16-значного названия
def generate_unique_filename(upload_folder, extension):
    while True:
        # Генерируем случайную строку из букв и цифр
        filename = ''.join(random.choices(string.ascii_letters + string.digits, k=16)) + f".{extension}"
        filepath = os.path.join(upload_folder, filename)
        
        # Проверяем, существует ли файл
        if not os.path.exists(filepath):
            return filename


def is_admin_wraps(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Проверка, является ли пользователь администратором
        if not (current_user.is_authenticated and current_user.is_active and int(current_user.level) >= 10):
            flash('Недостаточно прав для доступа к этой странице.', 'danger')
            return redirect(url_for('admin'))  # Перенаправление на главную страницу
        return f(*args, **kwargs)
    return decorated_function


def is_admin():
    return current_user.is_authenticated and current_user.is_active and int(current_user.level) >= 10


def get_admin_header():
    if is_admin():
        applications = Contract.query.filter(Contract.status == "checked").all()
        if len(applications) > 0:
            checked_len = f" ({len(applications)})"
        else:
            checked_len = ""
        menu_items = [
            {'url': '/new-expense', 'label': 'Новый расход', 'page': 'new-expense'},
            {'url': '/weekly-results', 'label': 'Итоги недели', 'page': 'weekly-results'},
            {'url': '/application-admin', 'label': f'Заявки на проверке{checked_len}', 'page': 'application-admin'},
        ]
        return menu_items
    return []


@app.before_request
def handle_unknown_page():
    # Проверяем, что пользователь перешел на несуществующую страницу
    if not request.endpoint or request.endpoint not in app.view_functions:
        if current_user.is_authenticated:
            # Если пользователь авторизован, перенаправляем на /admin
            return redirect(url_for('admin'))
        else:
            # Если пользователь не авторизован, перенаправляем на /
            return redirect(url_for('index'))


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    # Если пользователь уже авторизован, перенаправляем его на страницу admin
    if current_user.is_authenticated:
        flash("Вы уже авторизованы. Перейдите на страницу администрирования.", "success")
        return redirect(url_for('admin'))
    
    form = LoginForm()
    if form.validate_on_submit():
        username = form.username.data
        password = form.password.data
        
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            flash('Вход успешен', 'success')
            return redirect(url_for('admin'))
        else:
            flash('Неверное имя пользователя или пароль', 'danger')
    
    return render_template(
        'login.html', 
        form=form
    )


@app.route('/admin')
@login_required
def admin():
    applications = Contract.query.filter(Contract.status == "active").all()

    username = current_user.first_name
    active_application = current_user.active_applications
    return render_template('main.html', 
        applications=applications, 
        active_application=active_application,
        menu_items=get_admin_header(), 
        active_page='home', 
        profile_name=username
    )


@app.route('/applications')
@login_required
def applications():
    applications = Contract.query.all()

    if not applications:
        applications = []
    else:
        applications = reversed(applications)
    
    username = current_user.first_name
    active_application = current_user.active_applications
    
    # Передаем данные в шаблон
    return render_template('applications.html', 
        menu_items=get_admin_header(), 
        active_application=active_application,
        active_page='applications', 
        profile_name=username,
        applications=applications
    )


@app.route('/finance')
@login_required
def finance():
    # Получение доходов
    contracts = Contract.query.filter(Contract.status == "closed").all()  # Только активные контракты
    income = sum(service.sum for contract in contracts for service in contract.services)  # Суммируем стоимость всех услуг

    # Получение расходов
    expenses = Expense.query.all()  # Все расходы
    expense = sum(expense.sum for expense in expenses)  # Суммируем все расходы

    result = income - expense  # Результат (доход - расход)

    # История транзакций (можно добавить дополнительные фильтры по пользователю или другим критериям)
    historys = []
    for contract in contracts:
        for service in contract.services:
            historys.append({
                "id": f"Заявка №{contract.id}",
                "date": contract.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "description": f"Услуга: {service.name} для контракта {contract.id}",
                "amount": service.sum,
                "type": "Доход",
                "url": f"{contract.id}-application"
            })
    
    for exp in expenses:
        historys.append({
            "id": f"Расход №{exp.id}",
            "date": exp.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "description": exp.name,
            "amount": exp.sum,
            "type": "Расход",
            "url": f"{exp.id}-expense"
        })
    
    if historys:
        historys = reversed(historys)
    else:
        historys = []

    # Пользовательская информация
    username = current_user.first_name
    active_application = current_user.active_applications

    return render_template('finance.html', 
        menu_items=get_admin_header(),
        admin=is_admin(),
        active_application=active_application,
        income=income, 
        expense=expense, 
        result=result, 
        historys=historys, 
        active_page='finance', 
        profile_name=username
    )


                           

@app.route('/new-expense')
@login_required
@is_admin_wraps
def new_expense():
    username = current_user.first_name
    active_application = current_user.active_applications
    return render_template('newExpense.html', 
        menu_items=get_admin_header(), 
        active_application=active_application,
        active_page='new-expense', 
        profile_name=username
    )


@app.route('/new-application', methods=['GET', 'POST'])
@login_required
def new_application():
    if request.method == 'POST':
        description = request.form.get('expenseDescription')
        address = request.form.get('expenseTitle')
        client = request.form.get('expenseClient')
        number = request.form.get('expenseNumber')
        performer = f"{current_user.first_name} {current_user.last_name}"

        new_contract = Contract(
            description=description,
            address=address,
            client=client,
            number=number,
            performer=performer
        )
        db.session.add(new_contract)
        db.session.commit()

        token = app.config['TOKEN']
        group_id = app.config['GROUP_ID']

        client_ = "Не указано" if not client else client
        applications_last_id = Contract.query.filter(
            Contract.status == "active"
        ).order_by(Contract.id.desc()).first()

        if applications_last_id:
            last_id = applications_last_id.id
        else:
            last_id = None

        message = f"""
Новый заказ - #{last_id}
-----------
Описание:
```{description}```
-----------
Адрес: 
{address}
-----------
Клиент:
{client_}
-----------
Номер для связи:
```{number}```
-----------
Создал ```{performer}```
Посмотреть:
https://repair-31.ru/{last_id}-application
        """

        url = f'https://api.telegram.org/bot{token}/sendMessage'
        
        # Параметры запроса
        params = {
            'chat_id': group_id,
            'text': message,
            'parse_mode': 'Markdown'
        }

        # Отправка POST-запроса
        response = requests.post(url, params=params)
        
        # Проверка ответа
        if response.status_code == 200:
            print("Сообщение успешно отправлено!")
        else:
            print(f"Ошибка: {response.status_code}, {response.text}")

        return redirect(url_for('applications'))
    username = current_user.first_name
    active_application = current_user.active_applications
    return render_template('newApplication.html', 
        menu_items=get_admin_header(), 
        active_application=active_application,
        profile_name=username
    )


@app.route('/end-application', methods=['POST'])
@login_required
def end_application():
    expenseId = request.form.get('expenseId')
    
    if int(current_user.active_applications) == int(expenseId):
        # Получаем данные из формы
        expenseCreatedAt = request.form.get('expenseCreatedAt')
        expenseDescription = request.form.get('expenseDescription')
        expenseClient = request.form.get('expenseClient')
        expenseAddress = request.form.get('expenseAddress')
        expenseAmount = request.form.get('expenseAmount')
        expenseRecommendation = request.form.get('expenseRecommendation')
        expensePerformer = request.form.get('expensePerformer')
        expenseAcceotanceDate = request.form.get('expenseAcceotanceDate')

        # Получаем данные об услугах
        serviceNames = request.form.getlist('serviceName[]')
        servicePrices = request.form.getlist('servicePrice[]')
        serviceQuantities = request.form.getlist('serviceQuantity[]')
        serviceTotals = request.form.getlist('serviceTotal[]')
        serviceWarranties = request.form.getlist('serviceWarranty[]')

        # Проверка обязательных полей для заявки
        missing_fields = []

        if not expenseCreatedAt:
            missing_fields.append('Дата подачи заявки')
        if not expenseDescription:
            missing_fields.append('Описание')
        if not expenseClient:
            missing_fields.append('Клиент')
        if not expenseAddress:
            missing_fields.append('Адрес')
        if not expenseAmount:
            missing_fields.append('Сумма')
        if not expenseRecommendation:
            missing_fields.append('Рекомендации')
        if not expensePerformer:
            missing_fields.append('Исполнитель')
        if not expenseAcceotanceDate:
            missing_fields.append('Дата исполнения')

        # Проверка полей об услугах
        for i, name in enumerate(serviceNames):
            if not name or not servicePrices[i] or not serviceQuantities[i] or not serviceTotals[i]:
                missing_fields.append(f"Услуга {i+1} не заполнена корректно")

        # Определяем словарь с ключами для фотографий и их колонками в базе данных
        photos = {
            'photo1': 'photo_receipt',
            'photo2': 'photo_document_face',
            'photo3': 'photo_document_back'
        }

        photo_paths = {}  # Используем словарь для хранения путей к фото

        # Проверяем, что все три фото загружены и валидны
        for photo_key, column_name in photos.items():
            photo = request.files.get(photo_key)
            if photo and allowed_file(photo.filename):
                # Получаем расширение файла
                extension = photo.filename.rsplit('.', 1)[1].lower()

                # Генерируем уникальное имя
                filename = generate_unique_filename(app.config['UPLOAD_FOLDER'], extension)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

                # Сохраняем файл
                photo.save(filepath)
                photo_paths[column_name] = filepath  # Сохраняем путь к фото в словарь
            else:
                # Если хотя бы одно фото не загружено или не валидное
                missing_fields.append(f'{column_name} должно быть загружено и иметь допустимый формат.')


        # Если не загружены все три фотки
        if len(photo_paths) != 3:
            missing_fields.append('Пожалуйста, загрузите все три фотографии: чек, лицевая и задняя сторона документа.')

        # Если есть ошибки, возвращаем пользователя на форму
        if missing_fields:
            flash(f"Пожалуйста, заполните следующие поля: {', '.join(missing_fields)}", 'danger')
            return redirect(url_for('end_application', id=expenseId))

        acceotance_date = datetime.datetime.now()

        # Сохраняем данные в базу данных
        application = Contract.query.filter_by(id=int(expenseId)).first()
        application.status = 'checked'
        application.amount = expenseAmount
        application.acceptance_date = acceotance_date
        application.recommendations = expenseRecommendation
        if application.client != expenseClient:
            application.client = expenseClient

        # Обновляем пути к фотографиям из словаря
        application.photo_receipt = photo_paths.get('photo_receipt')
        application.photo_document_face = photo_paths.get('photo_document_face')
        application.photo_document_back = photo_paths.get('photo_document_back')

        db.session.commit()

        # Обработка услуг
        for i in range(len(serviceNames)):
            service_name = serviceNames[i]
            service_price = float(servicePrices[i]) if servicePrices[i] else 0.0
            service_quantity = int(serviceQuantities[i]) if serviceQuantities[i] else 0
            service_total = float(serviceTotals[i]) if serviceTotals[i] else 0.0
            service_warranty = serviceWarranties[i]

            service = Service(
                contract_id=expenseId,
                name=service_name,
                price=service_price,
                quantity=service_quantity,
                sum=service_total,
                warranty=service_warranty
            )
            
            db.session.add(service)
        
        db.session.commit()

        # Снимаем активную заявку с пользователя
        current_user.active_applications = None
        db.session.commit()

        # Успешно завершено, перенаправляем
        flash('Заявка успешно завершена.', 'success')
        return redirect(url_for('applications'))
    else:
        flash("Вы не можете завершить эту заявку.", 'danger')
        return redirect(url_for('applications'))



@app.route('/flash-message', methods=['POST'])
def flash_message():
    message = request.form.get('message')
    if message:
        flash(message, 'danger')
        return jsonify({'status': 'success'}), 200
    return jsonify({'status': 'error'}), 400


@app.route('/edit-<id>-application')
@login_required
def edit_application(id):
    id_application = f"#{id}"
    description = "Description for application #" + id
    address = "Address for application #" + id
    client = "Client for application #" + id

    username = current_user.first_name
    active_application = current_user.active_applications
    return render_template(
        'newEditReadApplications.html', 
        menu_items=get_admin_header(),
        id_application=id_application,
        active_application=active_application,
        description=description, 
        address=address, 
        client=client, 
        profile_name=username
    )


@app.route('/<id>-application')
@login_required
def read_application(id):
    application = Contract.query.filter(Contract.id == id).first()

    if application is None:
        flash("Заявка не найдена.", 'danger')
        return redirect(url_for('applications'))

    if application.status.value == "closed" or application.status.value == "checked":
        id_application = int(id)
        created_at = application.created_at.strftime("%Y-%m-%d %H:%M:%S")
        acceptance_date = application.acceptance_date.strftime("%Y-%m-%d %H:%M:%S")
        client = application.client
        description = application.description
        address = application.address
        number = application.number
        recommendation = application.recommendations
        performer = application.performer
        amount = application.amount

        photo1_db = application.photo_receipt
        photo2_db = application.photo_document_face
        photo3_db = application.photo_document_back
        
        scan1_db = application.scan_receipt
        scan2_db = application.scan_document_face
        scan3_db = application.scan_document_back

        if scan1_db and scan2_db and scan3_db:
            photo1 = str(scan1_db).replace("app/", "").replace("\\", "/")
            photo2 = str(scan2_db).replace("app/", "").replace("\\", "/")
            photo3 = str(scan3_db).replace("app/", "").replace("\\", "/")
        else:
            photo1 = str(photo1_db).replace("app/", "").replace("\\", "/")
            photo2 = str(photo2_db).replace("app/", "").replace("\\", "/")
            photo3 = str(photo3_db).replace("app/", "").replace("\\", "/")

        username = current_user.first_name
        active_application = current_user.active_applications

        return render_template('readFinishApplication.html', 
            id_application=id_application,
            created_at=created_at,
            acceptance_date=acceptance_date,
            status=application.status.value,
            client=client,
            description=description,
            address=address,
            number=number,
            recommendation=recommendation,
            performer=performer,
            amount=amount,
            active_application=active_application,
            photo1=photo1,
            photo2=photo2,
            photo3=photo3,
            profile_name=username,
            active_page='application',
            menu_items=get_admin_header(),
        )

    
    id_application = int(id)
    description = application.description
    address = application.address
    client = application.client
    number = application.number
    status = False if application.status.value == "closed" else True

    if current_user.active_applications == int(id):
        worked_application = True
    else:
        worked_application = False


    username = current_user.first_name
    active_application = current_user.active_applications
    checked = True if application.status.value == "checked" else False
    return render_template('readApplication.html', 
        active=status,
        worked=worked_application,
        checked=checked,
        active_application=active_application,
        closed=True if application.status.value == "closed" else False,
        menu_items=get_admin_header(),
        id_application=id_application,
        description=description,
        date = application.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        address=address, 
        client=client,
        number=number,
        profile_name=username,
        active_page='application'
    )


@app.route('/end-application-<int:id>')
@login_required
def end_application_id(id):
    if current_user.active_applications == id:
        application = Contract.query.filter(Contract.id == id).first()

        created_at = application.created_at.strftime("%Y-%m-%d %H:%M:%S")
        address = application.address
        client = application.client
        description = application.description
        performer = application.performer
        acceotance_date_pre = datetime.datetime.now()
        acceotance_date = acceotance_date_pre.strftime("%Y-%m-%d %H:%M:%S")


        username = current_user.first_name
        active_application = current_user.active_applications
        return render_template('endApplication.html', 
            menu_items=get_admin_header(),
            id=id,
            created_at=created_at,
            address=address,
            client=client,
            active_application=active_application,
            description=description,
            performer=performer,
            acceotance_date=acceotance_date,
            profile_name=username
        )
    flash("Вы не можете завершать эту заявку.", 'danger')
    return redirect(url_for('applications'))


@app.route('/delete_application/<int:id>', methods=['POST'])
@login_required
def delete_application(id):
    application = Contract.query.filter_by(id=id).first()

    if application:
        try:
            db.session.delete(application)
            db.session.commit()
            flash('Заявка успешно удалена!', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Ошибка при удалении заявки.', 'danger')
    else:
        flash('Заявка не найдена.', 'danger')

    return redirect(url_for('applications'))


@app.route('/apply_application/<int:id>', methods=['POST'])
@login_required
def apply_application(id):
    application = Contract.query.filter_by(id=id).first()

    if current_user.active_applications:
        flash('Вы уже в заявке.', 'danger')
        return redirect(url_for('admin'))
    
    users = User.query.filter_by(active_applications=id).all()
    if users:
        flash('Заявка занята', 'danger')
        return redirect(url_for('applications'))

    if application:
        try:
            # Изменяем статус заявки на 'worked'
            application.status = 'worked'
            application.performer = f"{current_user.first_name} {current_user.last_name}"
            db.session.commit()
            current_user.active_applications = id
            db.session.commit()
            flash('Заявка успешно принята!', 'success')  # Выводим сообщение об успешном принятии заяв
        except Exception as e:
            db.session.rollback()  # В случае ошибки откатываем изменения
            flash('Ошибка при принятии заявки.', 'danger')  # Выводим сообщение об ошибке при принятии заявки
    else:
        flash('Заявка не найдена.', 'danger')  # Выводим сообщение об отсутствии заявки

    return redirect(url_for('admin'))


@app.route('/application-admin')
@login_required
@is_admin_wraps
def application_admin():
    applications = Contract.query.filter(Contract.status == "checked").all()

    if not applications:
        applications = []
    else:
        applications = reversed(applications)
    
    username = current_user.first_name
    active_application = current_user.active_applications
    
    # Передаем данные в шаблон
    return render_template('applicationAdmin.html', 
        menu_items=get_admin_header(), 
        active_application=active_application,
        active_page='application-admin', 
        profile_name=username,
        applications=applications
    )



@app.route('/checked-<id>-application')
@login_required
@is_admin_wraps
def checked_read_application(id):
    application = Contract.query.filter(Contract.id == id).first()

    # Извлекаем все услуги для текущей заявки
    services = application.services

    # Преобразуем дату в нужный формат
    id_application = int(id)
    created_at = application.created_at.strftime("%Y-%m-%d %H:%M:%S")
    acceptance_date = application.acceptance_date.strftime("%Y-%m-%d %H:%M:%S")
    client = application.client
    description = application.description
    address = application.address
    number = application.number
    recommendation = application.recommendations
    performer = application.performer
    amount = application.amount

    photo1_db = application.photo_receipt
    photo2_db = application.photo_document_face
    photo3_db = application.photo_document_back

    photo1 = str(photo1_db).replace("app/", "").replace("\\", "/")
    photo2 = str(photo2_db).replace("app/", "").replace("\\", "/")
    photo3 = str(photo3_db).replace("app/", "").replace("\\", "/")

    username = current_user.first_name
    active_application = current_user.active_applications

    return render_template('readApplicationAdmin.html', 
        id_application=id_application,
        created_at=created_at,
        acceptance_date=acceptance_date,
        status=application.status.value,
        client=client,
        description=description,
        address=address,
        number=number,
        recommendation=recommendation,
        performer=performer,
        amount=amount,
        active_application=active_application,
        photo1=photo1,
        photo2=photo2,
        photo3=photo3,
        profile_name=username,
        active_page='application',
        services=services,
        menu_items=get_admin_header(),
    )



@app.route('/end-checked-application', methods=['POST'])
@login_required
@is_admin_wraps
def end_checked_application():
    expenseId = request.form.get('expenseId')
    expenseClient = request.form.get('expenseClient')
    expenseDescription = request.form.get('expenseDescription')
    expenseAddress = request.form.get('expenseAddress')
    expenseNumber = request.form.get('expenseNumber')
    expenseRecommendation = request.form.get('expenseRecommendation')
    expensePerformer = request.form.get('expensePerformer')
    expenseAmount = request.form.get('expenseAmount')

    # Проверка обязательных полей для заявки
    missing_fields = []

    if not expenseClient:
        missing_fields.append('Клиент')
    if not expenseDescription:
        missing_fields.append('Описание')
    if not expenseAddress:
        missing_fields.append('Адрес')
    if not expenseNumber:
        missing_fields.append('Номер телефона')
    if not expenseRecommendation:
        missing_fields.append('Рекомендации')
    if not expensePerformer:
        missing_fields.append('Исполнитель')
    if not expenseAmount:
        missing_fields.append('Сумма')

    # Определяем словарь с ключами для фотографий и их колонками в базе данных
    photos = {
        'photo1': 'photo_receipt',
        'photo2': 'photo_document_face',
        'photo3': 'photo_document_back'
    }

    photo_paths = {}  # Используем словарь для хранения путей к фото

    # Проверяем, что все три фото загружены и валидны
    for photo_key, column_name in photos.items():
        photo = request.files.get(photo_key)
        if photo and allowed_file(photo.filename):
            # Получаем расширение файла
            extension = photo.filename.rsplit('.', 1)[1].lower()

            # Генерируем уникальное имя
            filename = generate_unique_filename(app.config['UPLOAD_FOLDER'], extension)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

            # Сохраняем файл
            photo.save(filepath)
            photo_paths[column_name] = filepath  # Сохраняем путь к фото в словарь
        else:
            # Если хотя бы одно фото не загружено или не валидное
            missing_fields.append(f'{column_name} должно быть загружено и иметь допустимый формат.')


    # Если не загружены все три фотки
    if len(photo_paths) != 3:
        missing_fields.append('Пожалуйста, загрузите все три фотографии: чек, лицевая и задняя сторона документа.')

    # Если есть ошибки, возвращаем пользователя на форму
    if missing_fields:
        flash(f"Пожалуйста, заполните следующие поля: {', '.join(missing_fields)}", 'danger')
        return redirect(url_for('checked_read_application', id=expenseId))

    # Сохраняем данные в базу данных
    application = Contract.query.filter_by(id=int(expenseId)).first()

    if application.client != expenseClient:
        application.client = expenseClient
    if application.description != expenseDescription:
        application.description = expenseDescription
    if application.address != expenseAddress:
        application.address = expenseAddress
    if application.number != expenseNumber:
        application.number = expenseNumber
    if application.recommendations != expenseRecommendation:
        application.recommendations = expenseRecommendation
    if application.performer != expensePerformer:
        application.performer = expensePerformer
    if application.amount != expenseAmount:
        application.amount = expenseAmount

    application.status = 'closed'

    # Обновляем пути к фотографиям из словаря
    application.scan_receipt = photo_paths.get('photo_receipt')
    application.scan_document_face = photo_paths.get('photo_document_face')
    application.scan_document_back = photo_paths.get('photo_document_back')

    db.session.commit()

    # Снимаем активную заявку с пользователя
    current_user.active_applications = None
    db.session.commit()

    # Успешно завершено, перенаправляем
    flash('Заявка успешно закрыта.', 'success')
    return redirect(url_for('application_admin'))


@app.route('/new-expense-admin', methods=['POST'])
@login_required
@is_admin_wraps
def new_expense_admin():
    expenseTitle = request.form.get('expenseTitle')
    expenseDescription = request.form.get('expenseDescription')
    expenseAmount = request.form.get('expenseAmount')
    expenseDate = request.form.get('expenseDate')

    # Проверка обязательных полей для заявки
    missing_fields = []

    if not expenseTitle:
        missing_fields.append('Название')
    if not expenseDescription:
        missing_fields.append('Описание')
    if not expenseAmount:
        missing_fields.append('Сумма')
    if not expenseDate:
        missing_fields.append('Дата и время покупки')

    photos = {
        'photo1': 'photo_receipt'
    }

    photo_paths = {}

    for photo_key, column_name in photos.items():
        photo = request.files.get(photo_key)
        if photo and allowed_file(photo.filename):
            # Получаем расширение файла
            extension = photo.filename.rsplit('.', 1)[1].lower()
            
            # Генерируем уникальное имя
            filename = generate_unique_filename(app.config['UPLOAD_FOLDER'], extension)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Сохраняем файл
            photo.save(filepath)
            photo_paths[column_name] = filepath
        else:
            missing_fields.append(f'{column_name} должно быть загружено и иметь допустимый формат.')

    if len(photo_paths) != 1:
        missing_fields.append('Пожалуйста, загрузите чек от покупки.')

    # Если есть ошибки, возвращаем пользователя на форму
    if missing_fields:
        flash(f"Пожалуйста, заполните следующие поля: {', '.join(missing_fields)}", 'danger')
        return redirect(url_for('new_expense'))

    new_contract = Expense(
        name=expenseTitle,
        description=expenseDescription,
        performer=f"{current_user.first_name} {current_user.last_name}",
        sum=expenseAmount,
        scan_receipt=photo_paths.get('photo_receipt')
    )
    db.session.add(new_contract)
    db.session.commit()

    # Успешно завершено, перенаправляем
    flash('Трата добавлена', 'success')
    return redirect(url_for('application_admin'))


@app.route('/<id>-expense')
@login_required
def read_expense(id):
    expense = Expense.query.filter(Expense.id == id).first()

    if expense is None:
        flash('Расход не найден.', 'danger')
        return redirect(url_for('application_admin'))

    name=expense.name
    description=expense.description
    sum=expense.sum
    date=expense.created_at

    scan_db = expense.scan_receipt
    scan = str(scan_db).replace("app/", "").replace("\\", "/")

    username = current_user.first_name
    active_application = current_user.active_applications

    return render_template('readExpense.html', 
        id_expense=id,
        name=name,
        description=description,
        sum=sum,
        date=date,
        photo=scan,
        active_application=active_application,
        profile_name=username,
        active_page='application',
        menu_items=get_admin_header(),
    )


@app.route('/delete_expense/<int:id>', methods=['POST'])
@login_required
def delete_expense(id):
    expense = Expense.query.filter_by(id=id).first()

    if expense:
        try:
            db.session.delete(expense)
            db.session.commit()
            flash('Расход успешно удалена!', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Ошибка при удалении расхода.', 'danger')
    else:
        flash('Расход не найдена.', 'danger')

    return redirect(url_for('finance'))


@app.route('/profile')
@login_required
def profile():
    # Фильтруем заявки, выполненные пользователем
    contracts = Contract.query.filter(
        (Contract.performer == f"{current_user.first_name} {current_user.last_name}") &
        (Contract.status == "closed")  # Только завершенные контракты
    ).join(Service).all()

    # Считаем количество заявок
    total_requests = len(contracts)

    # Считаем количество услуг и общую сумму заработка
    total_services = 0
    total_earnings = 0.0
    for contract in contracts:
        for service in contract.services:
            total_services += service.quantity
            total_earnings += service.sum

    # Подготовим данные для отображения в шаблоне
    username = current_user.username
    first_name = current_user.first_name
    last_name = current_user.last_name
    active_application = current_user.active_applications

    return render_template('profile.html',
        first_name=first_name,
        last_name=last_name,
        username=username,
        profile_name=first_name,
        active_application=active_application,
        total_requests=total_requests,
        total_services=total_services,
        total_earnings=total_earnings,
        menu_items=get_admin_header(),
        active_page='profile'
    )



# Функция для завершения недели
def finish_week():
    # Создание новой папки с текущей датой
    folder_name = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    folder_path = os.path.join(app.config['WEEKLY_FOLDER'], folder_name)
    os.makedirs(folder_path)

    # Архив для фото
    zip_path = os.path.join(folder_path, "photos.zip")

    # Архивируем фотографии (фото из contracts и expenses)
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        # Добавляем фото из контрактов
        applications = Contract.query.filter_by(status='closed').all()
        add_photos_to_zip(zipf, applications, "contract")
        # Добавляем фото из расходов
        expenses = Expense.query.all()
        add_photos_to_zip(zipf, expenses, "expense")
    
    # Данные для Excel
    excel_data = create_excel_file()
    # Сохраняем Excel файл
    excel_path = os.path.join(folder_path, "week_data.xlsx")
    excel_data.save(excel_path)

    return folder_path


# Функция для создания Excel файла с несколькими листами
def create_excel_file():
    wb = Workbook()

    # Лист для контрактов
    ws_contracts = wb.create_sheet("Contracts")
    # Заголовки для Contracts
    headers_contracts = ["ID", "Created at", "Address", "Client", "Description", "Amount", "Recommendations", "Performer", "Number", "Acceptance date", "Scan receipt", "Scan face", "Scan back"]
    ws_contracts.append(headers_contracts)

    # Собирать данные из Contract
    contracts = Contract.query.filter_by(status='closed').all()
    for contract in contracts:
        ws_contracts.append([contract.id, contract.created_at, contract.address, contract.client, contract.description, 
                             contract.amount, contract.recommendations, contract.performer, contract.number, 
                             contract.acceptance_date, contract.scan_receipt, contract.scan_document_face, contract.scan_document_back])
    
    # Отключаем автоматический флашинг, чтобы избежать ошибок
    with db.session.no_autoflush:
        # Удаляем контракты и связанные с ними услуги
        for contract in contracts:
            # Удаляем связанные услуги
            for service in contract.services:
                db.session.delete(service)
            db.session.delete(contract)

        # Сохраняем изменения в базе данных
        db.session.commit()

    # Лист для услуг (Services)
    ws_services = wb.create_sheet("Services")
    # Заголовки для Services
    headers_services = ["ID", "Contract ID", "Name", "Price", "Quantity", "Sum", "Warranty"]
    ws_services.append(headers_services)

    # Собирать данные из Service
    for contract in contracts:
        for service in contract.services:
            ws_services.append([service.id, service.contract_id, service.name, service.price, service.quantity, service.sum, service.warranty])

    # Лист для расходов (Expenses)
    ws_expenses = wb.create_sheet("Expenses")
    # Заголовки для Expenses
    headers_expenses = ["ID", "Name", "Description", "Performer", "Sum", "Created at", "Scan receipt"]
    ws_expenses.append(headers_expenses)

    # Собирать данные из Expenses
    expenses = Expense.query.all()
    for expense in expenses:
        ws_expenses.append([expense.id, expense.name, expense.description, expense.performer, expense.sum, expense.created_at, expense.scan_receipt])
    
    # Удаляем расходы
    with db.session.no_autoflush:
        for expense in expenses:
            db.session.delete(expense)

        # Сохраняем изменения в базе данных
        db.session.commit()

    # Удаляем пустую вкладку, которая создается по умолчанию
    del wb['Sheet']

    return wb


# Функция для добавления фотографий в архив
def add_photos_to_zip(zipf, items, item_type):
    for item in items:
        # Пример фото: предполагаем, что фото находятся по этим путям
        if item_type == "contract":
            photo_paths = [item.photo_receipt, item.photo_document_face, item.photo_document_back, item.scan_receipt, item.scan_document_face, item.scan_document_back]
        elif item_type == "expense":
            photo_paths = [item.scan_receipt]
        for photo in photo_paths:
            if photo and os.path.exists(photo):
                zipf.write(photo, os.path.basename(photo))
        for photo_path in photo_paths:
            os.remove(photo_path)


@app.route('/finish_week', methods=['POST'])
@login_required
@is_admin_wraps
def finish_week_route():
    # Проверяем, есть ли хотя бы один контракт или хотя бы одна трата
    if not Contract.query.filter_by(status='closed').first() and not Expense.query.first():
        flash("Невозможно завершить неделю: нет завершенных контрактов или расходов.", 'danger')
        return redirect(url_for('weekly_results'))  # Перенаправляем на нужную страницу (например, страницу с ошибкой или главную)

    # Если контракты или расходы есть, продолжаем выполнение
    folder_path = finish_week()
    flash(f"Результаты недели сохранены в папке: {folder_path}", "success")
    return weekly_results() 


@app.route('/weekly-results')
@login_required
@is_admin_wraps
def weekly_results():
    # Получаем все папки в директории, сортируем их по имени (по дате)
    weekly_folder = app.config['WEEKLY_FOLDER']
    week_folders = [folder for folder in os.listdir(weekly_folder) if os.path.isdir(os.path.join(weekly_folder, folder))]
    week_folders.sort(reverse=True)

    first_name = current_user.first_name

    return render_template('weekly_results.html',
        week_folders=week_folders,
        profile_name=first_name,
        menu_items=get_admin_header(),
        active_page='weekly-results'
    )


@app.route('/download-weekly-<week_folder>')
@login_required
@is_admin_wraps
def download_weekly(week_folder):
    try:
        weekly_folder = app.config['WEEKLY_FOLDER']
        week_path = os.path.join(weekly_folder, week_folder).replace("/", "\\")

        # Проверяем, существует ли папка
        if not os.path.isdir(week_path):
            flash("Папка не найдена", 'danger')
            return redirect('weekly-results')

        # Фиксированные имена файлов
        archive_file = 'photos.zip'
        excel_file = 'week_data.xlsx'

        status_archive = os.path.exists(os.path.join(week_path, archive_file))
        status_excel = os.path.exists(os.path.join(week_path, excel_file))

        if not status_archive and not status_excel:
            flash("Не найдены архивный или Excel файлы в папке", 'danger')
            return redirect('weekly-results')

        # Создание временного архива в системе
        with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
            zip_filename = f"{tmp_file.name}.zip"

        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            if status_archive:
                zipf.write(os.path.join(week_path, archive_file), archive_file)
            if status_excel:
                zipf.write(os.path.join(week_path, excel_file), excel_file)

        # Отправляем архив с файлами
        return send_file(zip_filename, as_attachment=True)

    except Exception as e:
        print(e)
        flash("Произошла ошибка при скачивании файлов", 'danger')
        return redirect('weekly-results')




@app.route('/logout')
def logout():
    logout_user()
    flash('Вы успешно вышли', 'success')
    return redirect(url_for('login'))
